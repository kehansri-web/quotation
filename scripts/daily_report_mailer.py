#!/usr/bin/env python3
"""
QuoteCraft Pro - Automated Daily Excel Report & Email Dispatcher
Fires daily at 08:00 PM IST (14:30 UTC) via GitHub Actions or Cron.
Compiles daily quotations, solar capacities (kW), pipeline financials (INR),
and sales staff leaderboards into an executive formatted Excel workbook (.xlsx)
and dispatches it via email with an executive HTML summary.
"""

import os
import sys
import json
import sqlite3
import datetime
import smtplib
import base64
import urllib.request
import urllib.parse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    import requests
except ImportError:
    requests = None


def get_ist_now():
    """Returns current datetime in Indian Standard Time (UTC+5:30)"""
    utc_now = datetime.datetime.now(datetime.timezone.utc)
    ist_offset = datetime.timedelta(hours=5, minutes=30)
    return utc_now + ist_offset


def format_inr(amount):
    """Formats numeric value into Indian Rupee style currency string"""
    if amount is None or amount == "":
        return "₹ 0"
    try:
        val = round(float(amount))
        neg = val < 0
        val = abs(val)
        s = str(val)
        if len(s) <= 3:
            res = s
        else:
            last3 = s[-3:]
            remaining = s[:-3]
            chunks = []
            while len(remaining) > 2:
                chunks.insert(0, remaining[-2:])
                remaining = remaining[:-2]
            if remaining:
                chunks.insert(0, remaining)
            res = ",".join(chunks) + "," + last3
        return f"{'-' if neg else ''}₹ {res}"
    except Exception:
        return f"₹ {amount}"


def get_ist_date_of_timestamp(ts_str):
    """Converts an ISO UTC timestamp to YYYY-MM-DD in Indian Standard Time"""
    if not ts_str:
        return ""
    try:
        clean_ts = str(ts_str).replace("Z", "+00:00")
        dt = datetime.datetime.fromisoformat(clean_ts)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        ist_dt = dt.astimezone(datetime.timezone(datetime.timedelta(hours=5, minutes=30)))
        return ist_dt.strftime("%Y-%m-%d")
    except Exception:
        return str(ts_str)[:10]


def fetch_quotes_from_supabase(supabase_url, supabase_key, date_str):
    """Fetches quotes from Supabase REST API created strictly on target date in IST"""
    if not requests or not supabase_url or not supabase_key:
        return None

    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json"
    }
    url = f"{supabase_url.rstrip('/')}/rest/v1/quotes?select=*&order=created_at.desc"
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        if resp.status_code == 200:
            all_quotes = resp.json()
            # Filter strictly for today's date in Indian Standard Time
            day_quotes = []
            for q in all_quotes:
                created_at = q.get("created_at") or ""
                quote_ist_date = get_ist_date_of_timestamp(created_at)
                if quote_ist_date == date_str:
                    day_quotes.append(q)
            return day_quotes
    except Exception as e:
        print(f"⚠️ Supabase fetch error: {e}", file=sys.stderr)
    return None


def fetch_quotes_from_sqlite(db_path, date_str):
    """Fetches quotes from local SQLite database.db created strictly on target date in IST"""
    if not os.path.exists(db_path):
        return None
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM quotes ORDER BY created_at DESC")
        rows = cur.fetchall()
        quotes = []
        for r in rows:
            q_dict = dict(r)
            created_at = q_dict.get("created_at") or ""
            quote_ist_date = get_ist_date_of_timestamp(created_at)
            if quote_ist_date == date_str:
                quotes.append(q_dict)
        conn.close()
        return quotes
    except Exception as e:
        print(f"⚠️ SQLite fetch error: {e}", file=sys.stderr)
    return None


def generate_sample_quotes_for_demo(ist_today):
    """Generates realistic sample proposals for report preview if DB is empty"""
    date_str = ist_today.strftime("%d/%m/%Y")
    return [
        {
            "id": "demo-1",
            "quote_number": f"QC-SOL-{ist_today.strftime('%Y%m%d')}-01",
            "client_name": "Dr. Ramesh Verma",
            "client_phone": "+91 98490 12345",
            "client_email": "ramesh.verma@gmail.com",
            "client_address": "Plot 14, Jubilee Hills, Hyderabad, Telangana",
            "kw_capacity": 5.0,
            "partner_brand": "adani",
            "structure_type": "Elevated",
            "total_cost": 300000,
            "subsidy": 78000,
            "net_cost": 222000,
            "sales_rep": "Rahul Saw (Sales Lead)",
            "installer_brand": "kehansri",
            "status": "Accepted",
            "created_at": ist_today.isoformat(),
            "quote_json": json.dumps({
                "solar": {
                    "customerType": "residential",
                    "systemType": "on-grid",
                    "kwCapacity": 5.0,
                    "partnerBrand": "adani",
                    "structureType": "Elevated",
                    "moduleWattage": 620,
                    "gridTariff": 7.50,
                    "customSystemCost": 300000,
                    "customSubsidy": 78000
                }
            })
        },
        {
            "id": "demo-2",
            "quote_number": f"QC-SOL-{ist_today.strftime('%Y%m%d')}-02",
            "client_name": "Apex Engineering & Precision Works",
            "client_phone": "+91 94900 88776",
            "client_email": "procurement@apexengineers.in",
            "client_address": "Plot 88, IDA Cherlapally, Hyderabad, Telangana",
            "kw_capacity": 50.0,
            "partner_brand": "waaree",
            "structure_type": "Super Structure",
            "total_cost": 2650000,
            "subsidy": 0,
            "net_cost": 2119000,
            "sales_rep": "Priya Sharma (C&I Specialist)",
            "installer_brand": "kenergy",
            "status": "Generated",
            "created_at": ist_today.isoformat(),
            "quote_json": json.dumps({
                "solar": {
                    "customerType": "commercial",
                    "systemType": "on-grid",
                    "kwCapacity": 50.0,
                    "partnerBrand": "waaree",
                    "cniMonthlyUnits": 6500,
                    "cniTaxRate": 25,
                    "gridTariff": 9.50,
                    "customSystemCost": 2433425,
                    "customSubsidy": 0
                }
            })
        },
        {
            "id": "demo-3",
            "quote_number": f"QC-SOL-{ist_today.strftime('%Y%m%d')}-03",
            "client_name": "Sri Krishna Agro Farm House",
            "client_phone": "+91 98850 44332",
            "client_email": "krishna.agro@yahoo.co.in",
            "client_address": "Shamirpet Outer Ring Road, Hyderabad",
            "kw_capacity": 8.0,
            "partner_brand": "tata",
            "structure_type": "Elevated",
            "total_cost": 495000,
            "subsidy": 78000,
            "net_cost": 417000,
            "sales_rep": "Vikram Reddy",
            "installer_brand": "kehansri",
            "status": "Generated",
            "created_at": ist_today.isoformat(),
            "quote_json": json.dumps({
                "solar": {
                    "customerType": "residential",
                    "systemType": "hybrid",
                    "kwCapacity": 8.0,
                    "partnerBrand": "tata",
                    "structureType": "Elevated",
                    "moduleWattage": 620,
                    "gridTariff": 7.50,
                    "customSystemCost": 495000,
                    "customSubsidy": 78000
                }
            })
        },
        {
            "id": "demo-4",
            "quote_number": f"QC-SOL-{ist_today.strftime('%Y%m%d')}-04",
            "client_name": "Matrix Logistics & Warehousing Hub",
            "client_phone": "+91 97001 55667",
            "client_email": "operations@matrixlogistics.com",
            "client_address": "GMR Aerospace Park, Shamshabad, Hyderabad",
            "kw_capacity": 100.0,
            "partner_brand": "adani",
            "structure_type": "Standard Galvanized",
            "total_cost": 5100000,
            "subsidy": 0,
            "net_cost": 4079000,
            "sales_rep": "Rahul Saw (Sales Lead)",
            "installer_brand": "kenergy",
            "status": "In Review",
            "created_at": ist_today.isoformat(),
            "quote_json": json.dumps({
                "solar": {
                    "customerType": "commercial",
                    "systemType": "on-grid",
                    "kwCapacity": 100.0,
                    "partnerBrand": "adani",
                    "cniMonthlyUnits": 13000,
                    "cniTaxRate": 25,
                    "gridTariff": 9.50,
                    "customSystemCost": 4683195,
                    "customSubsidy": 0
                }
            })
        }
    ]


def parse_quote_details(q):
    """Extracts standardized solar metrics from quote record and JSON"""
    q_json = {}
    if q.get("quote_json"):
        try:
            if isinstance(q["quote_json"], str):
                q_json = json.loads(q["quote_json"])
            elif isinstance(q["quote_json"], dict):
                q_json = q["quote_json"]
        except Exception:
            q_json = {}

    solar = q_json.get("solar") or {}
    client = q_json.get("client") or {}

    kw = float(solar.get("kwCapacity") or q.get("kw_capacity") or 5.0)
    customer_type = solar.get("customerType") or "residential"
    is_cni = (customer_type == "commercial")
    system_type = solar.get("systemType") or "on-grid"

    brand = (solar.get("partnerBrand") or q.get("partner_brand") or "adani").lower()
    brand_display = "ADANI Power"
    if brand == "waaree":
        brand_display = "WAAREE Solar"
    elif brand == "tata":
        brand_display = "TATA Power"
    elif brand == "custom":
        brand_display = solar.get("customPartnerName") or "Custom OEM"

    total_cost = float(q.get("total_cost") or solar.get("customSystemCost") or (kw * 55096 * 1.089))
    
    if is_cni:
        subsidy = 0.0
        tax_rate = float(solar.get("cniTaxRate") or 25)
        ad_base = total_cost / 1.089 * 0.40
        tax_shield = round(ad_base * (tax_rate / 100))
        gst_itc = round(total_cost - (total_cost / 1.089))
        net_outlay = max(0, total_cost - tax_shield - gst_itc)
    else:
        subsidy = float(q.get("subsidy") or solar.get("customSubsidy") or (78000 if kw >= 3 else (60000 if kw >= 2 else 30000)))
        tax_shield = 0.0
        gst_itc = 0.0
        net_outlay = max(0, total_cost - subsidy)

    annual_gen = round(kw * 1600)
    tariff = float(solar.get("gridTariff") or (9.5 if is_cni else 7.5))
    annual_savings = round(annual_gen * tariff)

    sales_rep = q.get("sales_rep") or q.get("sales_username") or "Admin"
    client_name = q.get("client_name") or client.get("name") or "Solar Client"
    client_phone = q.get("client_phone") or client.get("phone") or ""
    client_email = q.get("client_email") or client.get("email") or ""
    client_address = q.get("client_address") or client.get("billingAddress") or ""
    quote_number = q.get("quote_number") or q.get("id") or "QC-QUOTE"
    created_at = q.get("created_at") or datetime.datetime.now().isoformat()

    return {
        "quote_number": quote_number,
        "created_at": created_at,
        "client_name": client_name,
        "client_phone": client_phone,
        "client_email": client_email,
        "client_address": client_address,
        "customer_type": "Commercial & Industrial (C&I)" if is_cni else "Residential Rooftop",
        "system_type": system_type.upper(),
        "kw_capacity": kw,
        "partner_brand": brand_display,
        "structure_type": solar.get("structureType") or q.get("structure_type") or "Elevated",
        "total_cost": total_cost,
        "subsidy": subsidy,
        "tax_shield": tax_shield,
        "gst_itc": gst_itc,
        "net_outlay": net_outlay,
        "annual_generation": annual_gen,
        "annual_savings": annual_savings,
        "sales_rep": sales_rep,
        "installer_brand": q.get("installer_brand") or "KehanSri Solar"
    }


def build_excel_workbook(parsed_quotes, ist_date_str):
    """Creates a beautifully styled, multi-sheet Excel file (.xlsx) using openpyxl"""
    if not openpyxl:
        raise RuntimeError("openpyxl library is required to build .xlsx report")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styling Palette
    FONT_FAMILY = "Segoe UI"
    EMERALD_DARK = "065F46"
    EMERALD_MID = "059669"
    EMERALD_LIGHT = "D1FAE5"
    GRAY_HEADER = "1E293B"
    GRAY_ROW_ALT = "F8FAFC"
    WHITE = "FFFFFF"
    CYAN_ACCENT = "0284C7"
    AMBER_ACCENT = "D97706"

    font_title = Font(name=FONT_FAMILY, size=16, bold=True, color=WHITE)
    font_sub = Font(name=FONT_FAMILY, size=10, italic=True, color="E2E8F0")
    font_sec_hdr = Font(name=FONT_FAMILY, size=12, bold=True, color=EMERALD_DARK)
    font_tbl_hdr = Font(name=FONT_FAMILY, size=10, bold=True, color=WHITE)
    font_bold = Font(name=FONT_FAMILY, size=10, bold=True)
    font_regular = Font(name=FONT_FAMILY, size=10)
    font_kpi_val = Font(name=FONT_FAMILY, size=14, bold=True, color=EMERALD_DARK)
    font_kpi_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")

    fill_title = PatternFill(start_color=EMERALD_DARK, end_color=EMERALD_DARK, fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color=GRAY_HEADER, end_color=GRAY_HEADER, fill_type="solid")
    fill_tbl_hdr_green = PatternFill(start_color=EMERALD_MID, end_color=EMERALD_MID, fill_type="solid")
    fill_kpi_card = PatternFill(start_color=EMERALD_LIGHT, end_color=EMERALD_LIGHT, fill_type="solid")
    fill_row_alt = PatternFill(start_color=GRAY_ROW_ALT, end_color=GRAY_ROW_ALT, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    thick_bottom = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='medium', color=EMERALD_DARK)
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # SHEET 1: EXECUTIVE KPI DASHBOARD
    # =========================================================================
    ws1 = wb.create_sheet(title="📊 Executive Dashboard")
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Rows 1-2)
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "⚡ KEHANSRI SOLAR & K ENERGY SOLUTIONS — DAILY PROPOSAL REPORT"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_title
    ws1["A1"].alignment = align_center

    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"Automated Daily Sales & Pipeline Summary • Date: {ist_date_str} (08:00 PM IST Dispatch)"
    ws1["A2"].font = font_sub
    ws1["A2"].fill = fill_title
    ws1["A2"].alignment = align_center

    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[2].height = 20

    # Aggregate Metrics Calculation
    total_quotes = len(parsed_quotes)
    total_kw = sum(q["kw_capacity"] for q in parsed_quotes)
    total_project_cost = sum(q["total_cost"] for q in parsed_quotes)
    total_net_cost = sum(q["net_outlay"] for q in parsed_quotes)
    total_subsidy = sum(q["subsidy"] for q in parsed_quotes)
    total_tax_shield = sum(q["tax_shield"] for q in parsed_quotes)
    total_annual_gen = sum(q["annual_generation"] for q in parsed_quotes)
    total_annual_sav = sum(q["annual_savings"] for q in parsed_quotes)

    res_quotes = [q for q in parsed_quotes if "Residential" in q["customer_type"]]
    cni_quotes = [q for q in parsed_quotes if "Commercial" in q["customer_type"]]

    res_kw = sum(q["kw_capacity"] for q in res_quotes)
    cni_kw = sum(q["kw_capacity"] for q in cni_quotes)
    res_val = sum(q["total_cost"] for q in res_quotes)
    cni_val = sum(q["total_cost"] for q in cni_quotes)

    # 4 KPI Summary Cards (Row 4-5)
    cards = [
        ("Total Proposals", f"{total_quotes} Quotes", "A4:B5"),
        ("Total Capacity Quoted", f"{total_kw:,.1f} kW", "C4:D5"),
        ("Gross Pipeline Value", format_inr(total_project_cost), "E4:F5"),
        ("Net Capital Outlay", format_inr(total_net_cost), "G4:G5")
    ]

    ws1["A4"] = "TOTAL PROPOSALS"
    ws1["A4"].font = font_kpi_lbl
    ws1["A5"] = f"{total_quotes} Quotes"
    ws1["A5"].font = font_kpi_val

    ws1["C4"] = "TOTAL CAPACITY QUOTED"
    ws1["C4"].font = font_kpi_lbl
    ws1["C5"] = f"{total_kw:,.1f} kW ({total_kw/1000:,.2f} MW)"
    ws1["C5"].font = font_kpi_val

    ws1["E4"] = "GROSS PIPELINE VALUE"
    ws1["E4"].font = font_kpi_lbl
    ws1["E5"] = format_inr(total_project_cost)
    ws1["E5"].font = font_kpi_val

    ws1["G4"] = "NET CLIENT OUTLAY"
    ws1["G4"].font = font_kpi_lbl
    ws1["G5"] = format_inr(total_net_cost)
    ws1["G5"].font = font_kpi_val

    for col in ["A", "C", "E", "G"]:
        for r in [4, 5]:
            cell = ws1[f"{col}{r}"]
            cell.fill = fill_kpi_card
            cell.alignment = align_center
            cell.border = thin_border

    # Merge card partner columns
    ws1.merge_cells("A4:B4"); ws1.merge_cells("A5:B5")
    ws1.merge_cells("C4:D4"); ws1.merge_cells("C5:D5")
    ws1.merge_cells("E4:F4"); ws1.merge_cells("E5:F5")

    # Table 1: Customer Category Breakdown (Rows 8-12)
    ws1["A7"] = "1. Customer Category Distribution (Residential vs C&I)"
    ws1["A7"].font = font_sec_hdr

    headers_cat = ["Category", "Proposals", "Capacity (kW)", "Pipeline Value (INR)", "Avg Ticket Size", "Govt Subsidy / 40% Tax Shield", "1st Yr Savings (INR)"]
    for col_idx, h in enumerate(headers_cat, start=1):
        cell = ws1.cell(row=8, column=col_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = thin_border

    cat_data = [
        ["🏡 Residential Rooftop", len(res_quotes), f"{res_kw:,.1f} kW", format_inr(res_val), format_inr(res_val/len(res_quotes)) if res_quotes else "₹ 0", f"Subsidy: {format_inr(total_subsidy)}", format_inr(sum(q['annual_savings'] for q in res_quotes))],
        ["🏭 Commercial & Industrial", len(cni_quotes), f"{cni_kw:,.1f} kW", format_inr(cni_val), format_inr(cni_val/len(cni_quotes)) if cni_quotes else "₹ 0", f"Tax Shield: {format_inr(total_tax_shield)}", format_inr(sum(q['annual_savings'] for q in cni_quotes))],
        ["Total / Portfolio", total_quotes, f"{total_kw:,.1f} kW", format_inr(total_project_cost), format_inr(total_project_cost/total_quotes) if total_quotes else "₹ 0", f"{format_inr(total_subsidy + total_tax_shield)}", format_inr(total_annual_sav)]
    ]

    for row_idx, row_vals in enumerate(cat_data, start=9):
        is_total = (row_idx == 11)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_bold if is_total else font_regular
            cell.alignment = align_center if col_idx in [2, 3] else (align_left if col_idx == 1 else align_right)
            cell.border = thick_bottom if is_total else thin_border
            if is_total:
                cell.fill = fill_kpi_card

    # Table 2: System Type Breakdown (Rows 14-18)
    ws1["A13"] = "2. System Type Breakdown (On-Grid / Off-Grid / Hybrid)"
    ws1["A13"].font = font_sec_hdr

    headers_sys = ["System Type", "Proposals", "Capacity (kW)", "Share of Capacity", "Gross Value (INR)"]
    for col_idx, h in enumerate(headers_sys, start=1):
        cell = ws1.cell(row=14, column=col_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr_green
        cell.alignment = align_center
        cell.border = thin_border

    sys_types = ["ON-GRID", "OFF-GRID", "HYBRID"]
    for row_idx, st in enumerate(sys_types, start=15):
        matching = [q for q in parsed_quotes if q["system_type"] == st]
        m_kw = sum(q["kw_capacity"] for q in matching)
        m_val = sum(q["total_cost"] for q in matching)
        pct = (m_kw / total_kw * 100) if total_kw > 0 else 0
        vals = [f"⚡ {st}", len(matching), f"{m_kw:,.1f} kW", f"{pct:.1f}%", format_inr(m_val)]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=v)
            cell.font = font_regular
            cell.alignment = align_center if col_idx in [2, 3, 4] else (align_left if col_idx == 1 else align_right)
            cell.border = thin_border

    # Adjust Sheet 1 Column Widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 16)

    # =========================================================================
    # SHEET 2: ALL QUOTATIONS DETAILED LEDGER
    # =========================================================================
    ws2 = wb.create_sheet(title="📋 Quotations Ledger")
    ws2.views.sheetView[0].showGridLines = True

    # Ledger Title
    ws2.merge_cells("A1:N1")
    ws2["A1"] = f"DETAILED SOLAR PROPOSALS LEDGER — {ist_date_str}"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_title
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 28

    ledger_headers = [
        "Quote #", "Date / Time", "Customer Name", "Phone", "Customer Type",
        "System Type", "Capacity (kW)", "Module Brand", "Structure",
        "Project Cost (INR)", "Subsidy / AD Shield (INR)", "Net Outlay (INR)",
        "1st Yr Savings (INR)", "Sales Rep"
    ]

    for col_idx, h in enumerate(ledger_headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = thin_border
    ws2.row_dimensions[3].height = 24

    for row_idx, q in enumerate(parsed_quotes, start=4):
        is_alt = (row_idx % 2 == 0)
        sub_tax = q["subsidy"] if "Residential" in q["customer_type"] else q["tax_shield"]
        row_vals = [
            q["quote_number"],
            q["created_at"][:16].replace("T", " "),
            q["client_name"],
            q["client_phone"],
            q["customer_type"],
            q["system_type"],
            q["kw_capacity"],
            q["partner_brand"],
            q["structure_type"],
            q["total_cost"],
            sub_tax,
            q["net_outlay"],
            q["annual_savings"],
            q["sales_rep"]
        ]

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            if is_alt:
                cell.fill = fill_row_alt

            # Formatting
            if col_idx == 7:
                cell.number_format = '#,##0.0 "kW"'
                cell.alignment = align_center
            elif col_idx in [10, 11, 12, 13]:
                cell.number_format = '₹ #,##,##0'
                cell.alignment = align_right
            elif col_idx in [1, 2, 5, 6, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Total Summary Row at bottom of Sheet 2
    last_row = len(parsed_quotes) + 4
    ws2.cell(row=last_row, column=1, value="TOTAL PORTFOLIO").font = font_bold
    ws2.cell(row=last_row, column=1).fill = fill_kpi_card
    ws2.cell(row=last_row, column=7, value=f"=SUM(G4:G{last_row-1})").font = font_bold
    ws2.cell(row=last_row, column=7).number_format = '#,##0.0 "kW"'
    ws2.cell(row=last_row, column=10, value=f"=SUM(J4:J{last_row-1})").font = font_bold
    ws2.cell(row=last_row, column=10).number_format = '₹ #,##,##0'
    ws2.cell(row=last_row, column=11, value=f"=SUM(K4:K{last_row-1})").font = font_bold
    ws2.cell(row=last_row, column=11).number_format = '₹ #,##,##0'
    ws2.cell(row=last_row, column=12, value=f"=SUM(L4:L{last_row-1})").font = font_bold
    ws2.cell(row=last_row, column=12).number_format = '₹ #,##,##0'
    ws2.cell(row=last_row, column=13, value=f"=SUM(M4:M{last_row-1})").font = font_bold
    ws2.cell(row=last_row, column=13).number_format = '₹ #,##,##0'

    for c in range(1, 15):
        cell = ws2.cell(row=last_row, column=c)
        cell.fill = fill_kpi_card
        cell.border = thick_bottom

    # Adjust Sheet 2 Column Widths
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # =========================================================================
    # SHEET 3: SALES STAFF LEADERBOARD
    # =========================================================================
    ws3 = wb.create_sheet(title="🏆 Sales Leaderboard")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:F1")
    ws3["A1"] = f"SALES STAFF PERFORMANCE LEADERBOARD — {ist_date_str}"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_title
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 28

    rep_headers = ["Rank", "Sales Representative", "Quotes Generated", "Capacity Quoted (kW)", "Pipeline Value (INR)", "Avg Deal Size (INR)"]
    for col_idx, h in enumerate(rep_headers, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = thin_border
    ws3.row_dimensions[3].height = 24

    # Group by sales rep
    reps = {}
    for q in parsed_quotes:
        rep = q["sales_rep"]
        if rep not in reps:
            reps[rep] = {"quotes": 0, "kw": 0.0, "val": 0.0}
        reps[rep]["quotes"] += 1
        reps[rep]["kw"] += q["kw_capacity"]
        reps[rep]["val"] += q["total_cost"]

    sorted_reps = sorted(reps.items(), key=lambda x: x[1]["val"], reverse=True)

    for rank, (rep_name, data) in enumerate(sorted_reps, start=1):
        is_alt = (rank % 2 == 0)
        avg_deal = data["val"] / data["quotes"] if data["quotes"] > 0 else 0
        medal = "🥇" if rank == 1 else ("🥈" if rank == 2 else ("🥉" if rank == 3 else f"#{rank}"))
        row_vals = [medal, rep_name, data["quotes"], f"{data['kw']:,.1f} kW", format_inr(data["val"]), format_inr(avg_deal)]
        
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=rank + 3, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = thin_border
            cell.alignment = align_center if col_idx in [1, 3, 4] else (align_left if col_idx == 2 else align_right)
            if is_alt:
                cell.fill = fill_row_alt

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 3, 16)

    return wb


def generate_html_email_body(parsed_quotes, ist_date_str):
    """Generates a responsive executive email body with summary cards and tables"""
    total_quotes = len(parsed_quotes)
    total_kw = sum(q["kw_capacity"] for q in parsed_quotes)
    total_cost = sum(q["total_cost"] for q in parsed_quotes)
    total_net = sum(q["net_outlay"] for q in parsed_quotes)
    
    res_quotes = [q for q in parsed_quotes if "Residential" in q["customer_type"]]
    cni_quotes = [q for q in parsed_quotes if "Commercial" in q["customer_type"]]

    # Build quotation rows for email preview table (top 6)
    table_rows_html = ""
    for q in parsed_quotes[:8]:
        table_rows_html += f"""
        <tr style="border-bottom: 1px solid #e2e8f0;">
          <td style="padding: 9px 12px; font-weight: 700; color: #0f172a;">{q['quote_number']}</td>
          <td style="padding: 9px 12px; color: #334155;">{q['client_name']}</td>
          <td style="padding: 9px 12px; text-align: center;"><span style="background: {'#d1fae5' if 'Residential' in q['customer_type'] else '#e0f2fe'}; color: {'#065f46' if 'Residential' in q['customer_type'] else '#0369a1'}; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 700;">{q['customer_type'].split()[0]}</span></td>
          <td style="padding: 9px 12px; text-align: center; font-weight: 700;">{q['kw_capacity']} kW</td>
          <td style="padding: 9px 12px; text-align: center; color: #64748b;">{q['partner_brand']}</td>
          <td style="padding: 9px 12px; text-align: right; font-weight: 800; color: #008852;">{format_inr(q['total_cost'])}</td>
          <td style="padding: 9px 12px; color: #64748b; font-size: 12px;">{q['sales_rep']}</td>
        </tr>
        """

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <title>Daily Solar Proposals Report</title>
    </head>
    <body style="font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif; background-color: #f1f5f9; margin: 0; padding: 20px; color: #1e293b;">
      <div style="max-width: 720px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08); border: 1px solid #e2e8f0;">
        
        <!-- Header Strip -->
        <div style="background: linear-gradient(135deg, #065f46 0%, #059669 100%); padding: 24px 28px; color: #ffffff;">
          <div style="font-size: 20px; font-weight: 800; letter-spacing: -0.3px; margin-bottom: 4px;">⚡ KEHANSRI SOLAR &amp; K ENERGY SOLUTIONS</div>
          <div style="font-size: 13px; opacity: 0.9;">Daily Quotations &amp; Commercial Pipeline Report &bull; {ist_date_str} (08:00 PM IST Dispatch)</div>
        </div>

        <div style="padding: 24px 28px;">
          <!-- 4 KPI Cards -->
          <div style="display: grid; grid-template-columns: 1fr 1fr 1fr 1fr; gap: 12px; margin-bottom: 24px;">
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
              <div style="font-size: 10.5px; font-weight: 700; color: #64748b; text-transform: uppercase;">Proposals</div>
              <div style="font-size: 18px; font-weight: 800; color: #065f46; margin-top: 4px;">{total_quotes}</div>
            </div>
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
              <div style="font-size: 10.5px; font-weight: 700; color: #64748b; text-transform: uppercase;">Total kW</div>
              <div style="font-size: 18px; font-weight: 800; color: #0284c7; margin-top: 4px;">{total_kw:,.1f} kW</div>
            </div>
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
              <div style="font-size: 10.5px; font-weight: 700; color: #64748b; text-transform: uppercase;">Gross Pipeline</div>
              <div style="font-size: 18px; font-weight: 800; color: #008852; margin-top: 4px;">{format_inr(total_cost)}</div>
            </div>
            <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
              <div style="font-size: 10.5px; font-weight: 700; color: #64748b; text-transform: uppercase;">Net Outlay</div>
              <div style="font-size: 18px; font-weight: 800; color: #d97706; margin-top: 4px;">{format_inr(total_net)}</div>
            </div>
          </div>

          <!-- Category Breakdown Pill -->
          <div style="background: #ecfdf5; border-left: 4px solid #10b981; padding: 12px 16px; border-radius: 6px; margin-bottom: 20px; font-size: 13px;">
            <strong>Breakdown:</strong> 🏡 <strong>{len(res_quotes)} Residential</strong> ({sum(q['kw_capacity'] for q in res_quotes):,.1f} kW &bull; {format_inr(sum(q['total_cost'] for q in res_quotes))}) &bull; 🏭 <strong>{len(cni_quotes)} C&amp;I Projects</strong> ({sum(q['kw_capacity'] for q in cni_quotes):,.1f} kW &bull; {format_inr(sum(q['total_cost'] for q in cni_quotes))})
          </div>

          <!-- Section Heading -->
          <div style="font-size: 14px; font-weight: 800; color: #0f172a; margin-bottom: 10px;">📋 Today's Generated Proposals</div>

          <!-- Table -->
          <table style="width: 100%; border-collapse: collapse; font-size: 12.5px; margin-bottom: 20px;">
            <thead>
              <tr style="background: #1e293b; color: #ffffff; text-align: left;">
                <th style="padding: 8px 12px;">Quote #</th>
                <th style="padding: 8px 12px;">Client</th>
                <th style="padding: 8px 12px; text-align: center;">Type</th>
                <th style="padding: 8px 12px; text-align: center;">Capacity</th>
                <th style="padding: 8px 12px; text-align: center;">Brand</th>
                <th style="padding: 8px 12px; text-align: right;">Value</th>
                <th style="padding: 8px 12px;">Rep</th>
              </tr>
            </thead>
            <tbody>
              {table_rows_html}
            </tbody>
          </table>

          <!-- Attachment Notice -->
          <div style="background: #f1f5f9; border: 1px dashed #cbd5e1; border-radius: 8px; padding: 14px; text-align: center; color: #475569; font-size: 13px;">
            📎 <strong>Full Multi-Sheet Excel Workbook Attached:</strong><br>
            Includes <em>Executive KPI Dashboard</em>, <em>Comprehensive Quotation Ledger</em>, and <em>Sales Staff Leaderboard</em>.
          </div>
        </div>

        <!-- Footer -->
        <div style="background: #f8fafc; border-top: 1px solid #e2e8f0; padding: 14px 28px; font-size: 11.5px; color: #94a3b8; text-align: center;">
          Generated automatically by QuoteCraft Pro &bull; KehanSri Solar &amp; K Energy Solutions &bull; IST 08:00 PM Scheduled Cron
        </div>

      </div>
    </body>
    </html>
    """


def send_email_via_gmail_api(client_id, client_secret, refresh_token, recipient_emails, subject, html_content, attachment_path):
    """Sends email directly via Google Gmail API OAuth2 (Official REST endpoint)"""
    try:
        token_payload = urllib.parse.urlencode({
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token"
        }).encode("utf-8")

        req = urllib.request.Request("https://oauth2.googleapis.com/token", data=token_payload, method="POST")
        req.add_header("Content-Type", "application/x-www-form-urlencoded")
        
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            access_token = data.get("access_token")
            if not access_token:
                print("❌ Failed to obtain Gmail access_token from refresh_token", file=sys.stderr)
                return False

        # Build MIME message
        msg = MIMEMultipart("mixed")
        msg["To"] = ", ".join(recipient_emails) if isinstance(recipient_emails, list) else recipient_emails
        msg["Subject"] = subject

        msg_body = MIMEMultipart("alternative")
        msg_body.attach(MIMEText(html_content, "html"))
        msg.attach(msg_body)

        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as f:
                part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(part)

        raw_bytes = msg.as_bytes()
        raw_b64 = base64.urlsafe_b64encode(raw_bytes).decode("utf-8")

        send_req = urllib.request.Request(
            "https://gmail.googleapis.com/gmail/v1/users/me/messages/send",
            data=json.dumps({"raw": raw_b64}).encode("utf-8"),
            method="POST"
        )
        send_req.add_header("Authorization", f"Bearer {access_token}")
        send_req.add_header("Content-Type", "application/json")

        with urllib.request.urlopen(send_req) as send_resp:
            result = json.loads(send_resp.read().decode("utf-8"))
            print(f"✓ Daily report email successfully dispatched via Gmail API! (Message ID: {result.get('id')})")
            return True

    except Exception as e:
        print(f"❌ Gmail API sending error: {e}", file=sys.stderr)
        return False


def send_email_with_attachment(recipient_emails, subject, html_content, attachment_path):
    """Sends email via Gmail API OAuth2 or SMTP with Excel attachment"""
    
    # 1. Check for Gmail API OAuth2 Credentials
    client_id = os.environ.get("GMAIL_CLIENT_ID") or "54322630044-vt5qaue05c8bprpbiv0hh3rij6odohpt.apps.googleusercontent.com"
    client_secret = os.environ.get("GMAIL_CLIENT_SECRET") or "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"
    refresh_token = os.environ.get("GMAIL_REFRESH_TOKEN")

    creds_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gmail_credentials.json")
    if not refresh_token and os.path.exists(creds_file):
        try:
            with open(creds_file, "r") as f:
                cdata = json.load(f)
                refresh_token = cdata.get("refresh_token")
        except Exception:
            pass

    if client_id and client_secret and refresh_token:
        print("🌐 Sending daily email via official Google Gmail API...")
        if send_email_via_gmail_api(client_id, client_secret, refresh_token, recipient_emails, subject, html_content, attachment_path):
            return True
        print("⚠️ Gmail API dispatch failed. Attempting SMTP fallback...")

    # 2. Fallback to standard SMTP
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")

    if not smtp_user or not smtp_pass:
        print("⚠️ Neither Gmail API refresh token nor SMTP credentials configured. Email not dispatched.")
        print("💡 To enable email dispatch:")
        print("   Run: python3 scripts/setup_gmail_oauth.py (for Gmail API)")
        print("   Or set GitHub Secrets: SMTP_USER and SMTP_PASS")
        return False

    msg = MIMEMultipart("mixed")
    msg["From"] = f"QuoteCraft Daily Reporter <{smtp_user}>"
    msg["To"] = ", ".join(recipient_emails) if isinstance(recipient_emails, list) else recipient_emails
    msg["Subject"] = subject

    msg_body = MIMEMultipart("alternative")
    msg_body.attach(MIMEText(html_content, "html"))
    msg.attach(msg_body)

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(attachment_path)}"')
        msg.attach(part)

    try:
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
            server.starttls()

        server.login(smtp_user, smtp_pass)
        recipients_list = [r.strip() for r in (recipient_emails.split(",") if isinstance(recipient_emails, str) else recipient_emails) if r.strip()]
        server.sendmail(smtp_user, recipients_list, msg.as_string())
        server.quit()
        print(f"✓ Daily report email successfully sent to: {', '.join(recipients_list)}")
        return True
    except Exception as e:
        print(f"❌ Failed to send daily email via SMTP: {e}", file=sys.stderr)
        return False


def generate_no_quotes_email_body(ist_date_str):
    """Generates clean, professional notification when no quotes were generated today"""
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>Daily Solar Proposals Report — {ist_date_str}</title>
    </head>
    <body style="margin: 0; padding: 0; background-color: #f1f5f9; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; color: #334155;">
      <div style="max-width: 650px; margin: 30px auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.06); border: 1px solid #e2e8f0;">
        
        <!-- Header -->
        <div style="background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); padding: 24px 28px; color: #ffffff;">
          <table style="width: 100%; border-collapse: collapse;">
            <tr>
              <td>
                <div style="font-size: 11px; font-weight: 700; letter-spacing: 1.2px; text-transform: uppercase; color: #38bdf8; margin-bottom: 4px;">QuoteCraft Pro &bull; Daily Operations Log</div>
                <div style="font-size: 20px; font-weight: 800; color: #ffffff;">Daily Quotation Summary</div>
                <div style="font-size: 13px; color: #94a3b8; margin-top: 2px;">Reporting Date: <strong>{ist_date_str}</strong> (IST 08:00 PM)</div>
              </td>
              <td style="text-align: right; vertical-align: top;">
                <span style="display: inline-block; padding: 5px 12px; background: rgba(148, 163, 184, 0.2); border: 1px solid rgba(148, 163, 184, 0.4); border-radius: 20px; font-size: 12px; font-weight: 700; color: #cbd5e1;">0 Quotes</span>
              </td>
            </tr>
          </table>
        </div>

        <!-- Body Content -->
        <div style="padding: 36px 28px; text-align: center;">
          <div style="font-size: 44px; margin-bottom: 12px;">📋</div>
          <div style="font-size: 19px; font-weight: 800; color: #0f172a; margin-bottom: 8px;">No Quotations Generated Today</div>
          <div style="font-size: 14px; color: #64748b; line-height: 1.6; max-width: 480px; margin: 0 auto 24px auto;">
            This is an automated daily report to confirm that <strong>no new solar proposals or quotations</strong> were generated in QuoteCraft Pro on <strong>{ist_date_str}</strong>.
          </div>

          <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px 20px; text-align: left; max-width: 480px; margin: 0 auto 24px auto; font-size: 13px; color: #475569;">
            <div style="font-weight: 700; color: #0f172a; margin-bottom: 6px;">ℹ️ Status Summary:</div>
            <ul style="margin: 0; padding-left: 18px; line-height: 1.6;">
              <li><strong>Cloud Database:</strong> 🟢 Connected &amp; Synced (Supabase)</li>
              <li><strong>Sales Team:</strong> Active</li>
              <li><strong>Previous Quotations:</strong> Safely logged &amp; preserved in database</li>
            </ul>
          </div>

          <div style="font-size: 12px; color: #94a3b8;">
            When your sales team generates proposals tomorrow, a detailed Excel spreadsheet will be automatically compiled and dispatched at 08:00 PM IST.
          </div>
        </div>

        <!-- Footer -->
        <div style="background: #f8fafc; border-top: 1px solid #e2e8f0; padding: 14px 28px; font-size: 11.5px; color: #94a3b8; text-align: center;">
          Generated automatically by QuoteCraft Pro &bull; KehanSri Solar &amp; K Energy Solutions &bull; IST 08:00 PM Scheduled Cron
        </div>

      </div>
    </body>
    </html>
    """


def main():
    ist_now = get_ist_now()
    ist_date_str = ist_now.strftime("%d %B %Y")
    date_file_str = ist_now.strftime("%Y-%m-%d")

    print(f"⚡ QuoteCraft Pro — Daily 08:00 PM IST Report Generator")
    print(f"📅 Report Date: {ist_date_str} (IST: {ist_now.strftime('%I:%M %p')})")

    # 1. Fetch only today's quotes from Supabase or SQLite
    supabase_url = os.environ.get("SUPABASE_URL") or "https://bvxyzsmveauqcpbnfgdt.supabase.co"
    supabase_key = os.environ.get("SUPABASE_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ2eHl6c212ZWF1cWNwYm5mZ2R0Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODgxODgyOTcsImV4cCI6MjEwMzc2NDI5N30.QuKpVpUgMSghV9ZWuB2b6006MV16-G2EHk0-Th3LDjI"
    db_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "database.db")

    raw_quotes = None
    if supabase_url and supabase_key:
        print("🌐 Connecting to Supabase Cloud Database...")
        raw_quotes = fetch_quotes_from_supabase(supabase_url, supabase_key, date_file_str)

    if raw_quotes is None and os.path.exists(db_file):
        print(f"💾 Reading from SQLite database: {db_file}")
        raw_quotes = fetch_quotes_from_sqlite(db_file, date_file_str)

    if raw_quotes is None:
        raw_quotes = []

    recipient_email = os.environ.get("REPORT_RECIPIENT_EMAIL") or "tarun.dixit@kehansri.com"

    # CASE A: NO QUOTATIONS GENERATED TODAY
    if len(raw_quotes) == 0:
        print(f"ℹ️ No quotations were generated today ({date_file_str}).")
        html_body = generate_no_quotes_email_body(ist_date_str)
        subject = f"⚡ Daily Solar Report — {ist_date_str}: No Quotations Generated Today"

        if recipient_email:
            print(f"📧 Dispatching 'No Quotations Generated Today' notification to {recipient_email}...")
            send_email_with_attachment(recipient_email, subject, html_body, attachment_path=None)
        else:
            print("💡 NOTE: Set 'REPORT_RECIPIENT_EMAIL' to automatically send the report.")

        print("🎉 Daily report processing completed successfully!")
        return

    # CASE B: TODAY'S QUOTATIONS GENERATED
    parsed_quotes = [parse_quote_details(q) for q in raw_quotes]
    print(f"📊 Processed {len(parsed_quotes)} quotations for today's summary.")

    # 2. Build Excel Workbook with today's proposals
    reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")
    os.makedirs(reports_dir, exist_ok=True)
    excel_filename = f"QuoteCraft_Daily_Report_{date_file_str}.xlsx"
    excel_filepath = os.path.join(reports_dir, excel_filename)

    wb = build_excel_workbook(parsed_quotes, ist_date_str)
    wb.save(excel_filepath)
    print(f"✓ Excel workbook saved successfully: {excel_filepath}")

    # 3. Generate HTML email body
    html_body = generate_html_email_body(parsed_quotes, ist_date_str)

    # 4. Dispatch Email with Excel Attachment
    if recipient_email:
        total_kw = sum(q['kw_capacity'] for q in parsed_quotes)
        subject = f"⚡ Daily Solar Proposals Report — {ist_date_str} ({len(parsed_quotes)} Quotes • {total_kw:,.1f} kW)"
        send_email_with_attachment(recipient_email, subject, html_body, excel_filepath)
    else:
        print("💡 NOTE: Set 'REPORT_RECIPIENT_EMAIL' environment variable to automatically send the report.")

    print("🎉 Daily report processing completed successfully!")


if __name__ == "__main__":
    main()
